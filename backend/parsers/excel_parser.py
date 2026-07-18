"""
Excel Parser — Reads existing Excel/ODS files and normalizes to internal Table model.
"""
from __future__ import annotations

import logging
import time
from pathlib import Path

from backend.models import Cell, Table, ProcessingResult, ProcessingStatus
from backend.utils import detect_data_type

logger = logging.getLogger(__name__)


class ExcelParser:
    """Parses Excel (.xlsx, .xls) and ODS files into internal Table model."""

    def parse(self, file_path: str | Path) -> ProcessingResult:
        """Parse an Excel or ODS file."""
        result = ProcessingResult()
        result.filename = Path(file_path).name
        result.file_category = "excel"
        start_time = time.time()

        ext = Path(file_path).suffix.lower()

        try:
            result.status = ProcessingStatus.OCR_IN_PROGRESS

            if ext in (".xlsx", ".xls"):
                tables = self._parse_openpyxl(str(file_path))
                result.ocr_engine_used = "openpyxl"
            elif ext == ".ods":
                tables = self._parse_ods(str(file_path))
                result.ocr_engine_used = "odfpy"
            else:
                result.mark_failed(f"Unsupported Excel format: {ext}")
                return result

            result.tables = tables
            result.page_count = len(tables)
            result.status = ProcessingStatus.READY_FOR_REVIEW
            result.processing_time = time.time() - start_time

        except Exception as e:
            result.mark_failed(f"Excel parsing failed: {str(e)}")
            logger.error(f"Excel parsing error: {e}", exc_info=True)

        return result

    def _parse_openpyxl(self, file_path: str) -> list[Table]:
        """Parse using openpyxl."""
        from openpyxl import load_workbook

        tables: list[Table] = []
        wb = load_workbook(file_path, read_only=True, data_only=True)

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                table = Table()
                table.source_engine = "openpyxl"
                table.title = sheet_name

                row_count = 0
                for row_idx, row in enumerate(ws.iter_rows(values_only=False)):
                    cells: list[Cell] = []
                    has_data = False

                    for col_idx, xl_cell in enumerate(row):
                        cell = Cell(row=row_idx, col=col_idx)
                        value = xl_cell.value

                        if value is not None:
                            has_data = True
                            cell.raw_value = str(value)
                            cell.value = value
                            cell.confidence = 1.0

                            # Preserve existing formatting
                            if xl_cell.font:
                                cell.font_bold = xl_cell.font.bold or False
                                if xl_cell.font.size:
                                    cell.font_size = xl_cell.font.size

                            if xl_cell.number_format and xl_cell.number_format != "General":
                                cell.format_string = xl_cell.number_format

                            # Detect data type for string values
                            if isinstance(value, str):
                                dt, parsed, fmt = detect_data_type(value)
                                cell.data_type = dt
                                if dt in ("number", "currency", "percentage"):
                                    cell.value = parsed
                                if not cell.format_string:
                                    cell.format_string = fmt
                            elif isinstance(value, (int, float)):
                                cell.data_type = "number"
                            elif hasattr(value, 'strftime'):
                                cell.data_type = "date"

                            # First row header detection
                            if row_idx == 0:
                                cell.is_header = True
                                cell.font_bold = True

                        cells.append(cell)

                    if has_data:
                        table.cells.append(cells)
                        row_count += 1

                if table.cells:
                    table.headers = [str(c.value) if c.value else "" for c in table.cells[0]]
                    table.num_rows = len(table.cells)
                    table.num_cols = max(len(r) for r in table.cells) if table.cells else 0
                    tables.append(table)

        finally:
            wb.close()

        return tables

    def _parse_ods(self, file_path: str) -> list[Table]:
        """Parse ODS files using odfpy or pandas."""
        try:
            import pandas as pd
            dfs = pd.read_excel(file_path, sheet_name=None, engine="odf")

            tables: list[Table] = []
            for sheet_name, df in dfs.items():
                table = self._dataframe_to_table(df, sheet_name)
                if table:
                    tables.append(table)

            return tables
        except ImportError:
            logger.warning("odfpy/pandas not available for ODS parsing")
            return []

    def _dataframe_to_table(self, df, sheet_name: str = "") -> Table | None:
        """Convert a pandas DataFrame to a Table."""
        if df.empty:
            return None

        table = Table()
        table.source_engine = "pandas"
        table.title = sheet_name

        # Header row
        header_cells: list[Cell] = []
        for col_idx, col_name in enumerate(df.columns):
            cell = Cell(row=0, col=col_idx)
            cell.value = str(col_name)
            cell.raw_value = str(col_name)
            cell.is_header = True
            cell.font_bold = True
            cell.confidence = 1.0
            header_cells.append(cell)
        table.cells.append(header_cells)
        table.headers = [str(c.value) for c in header_cells]

        # Data rows
        for row_idx, (_, row_data) in enumerate(df.iterrows()):
            cells: list[Cell] = []
            for col_idx, value in enumerate(row_data):
                cell = Cell(row=row_idx + 1, col=col_idx)
                if value is not None and str(value) != "nan":
                    cell.raw_value = str(value)
                    cell.value = value
                    cell.confidence = 1.0

                    if isinstance(value, str):
                        dt, parsed, fmt = detect_data_type(value)
                        cell.data_type = dt
                        if dt in ("number", "currency", "percentage"):
                            cell.value = parsed
                        cell.format_string = fmt
                    elif isinstance(value, (int, float)):
                        cell.data_type = "number"

                cells.append(cell)
            table.cells.append(cells)

        table.num_rows = len(table.cells)
        table.num_cols = len(df.columns)
        return table
