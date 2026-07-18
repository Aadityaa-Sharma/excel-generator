"""
Excel Generator — Produces professional, formatted Excel workbooks.
Uses openpyxl for .xlsx generation with:
- Auto-width columns
- Frozen headers
- Auto-filters
- Bold headings with color
- Currency, date, percentage formatting
- Low-confidence highlighting
- Validation error highlighting
- Named worksheets
- Merged cells
- CA-specific summary sheets
"""
from __future__ import annotations

import logging
import os
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

from backend.models import Cell, Table, ProcessingResult, ValidationResult
from config import Config

logger = logging.getLogger(__name__)


# ── Style Definitions ───────────────────────────────────────────────────────

class ExcelStyles:
    """Centralized style definitions for the Excel workbook."""

    def __init__(self):
        header_color = Config.EXCEL_HEADER_COLOR
        self.header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        self.header_font = Font(
            name=Config.EXCEL_DEFAULT_FONT,
            size=Config.EXCEL_DEFAULT_FONT_SIZE,
            bold=True, color="FFFFFF"
        )
        self.data_font = Font(
            name=Config.EXCEL_DEFAULT_FONT,
            size=Config.EXCEL_DEFAULT_FONT_SIZE,
        )
        self.bold_font = Font(
            name=Config.EXCEL_DEFAULT_FONT,
            size=Config.EXCEL_DEFAULT_FONT_SIZE,
            bold=True,
        )

        self.low_confidence_fill = PatternFill(
            start_color=Config.EXCEL_LOW_CONFIDENCE_COLOR,
            end_color=Config.EXCEL_LOW_CONFIDENCE_COLOR,
            fill_type="solid"
        )
        self.error_fill = PatternFill(
            start_color=Config.EXCEL_ERROR_COLOR,
            end_color=Config.EXCEL_ERROR_COLOR,
            fill_type="solid"
        )
        self.warning_fill = PatternFill(
            start_color="FFF3CD",
            end_color="FFF3CD",
            fill_type="solid"
        )
        self.info_fill = PatternFill(
            start_color="D1ECF1",
            end_color="D1ECF1",
            fill_type="solid"
        )
        self.total_fill = PatternFill(
            start_color="E8F5E9",
            end_color="E8F5E9",
            fill_type="solid"
        )

        thin_border = Side(border_style="thin", color="D0D0D0")
        self.cell_border = Border(
            left=thin_border, right=thin_border,
            top=thin_border, bottom=thin_border
        )
        header_border = Side(border_style="thin", color="FFFFFF")
        self.header_border = Border(
            left=header_border, right=header_border,
            top=header_border, bottom=header_border
        )

        self.center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)

        # Number formats
        self.currency_format_inr = '₹#,##0.00'
        self.currency_format_usd = '$#,##0.00'
        self.number_format = '#,##0.00'
        self.integer_format = '#,##0'
        self.percentage_format = '0.00%'
        self.date_format = 'DD/MM/YYYY'
        self.text_format = '@'


class ExcelGenerator:
    """
    Generates professional Excel workbooks from ProcessingResult.
    """

    def __init__(self):
        self.styles = ExcelStyles()

    def generate(self, result: ProcessingResult, output_path: str | Path | None = None,
                 include_summary: bool = True, include_validation: bool = True) -> str:
        """
        Generate an Excel workbook from processing results.
        Returns the path to the generated .xlsx file.
        """
        if not output_path:
            output_dir = Config.OUTPUT_DIR
            output_dir.mkdir(parents=True, exist_ok=True)
            base_name = Path(result.filename).stem if result.filename else "output"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = output_dir / f"{base_name}_{timestamp}.xlsx"

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Create data sheets
        for idx, table in enumerate(result.tables):
            sheet_name = self._safe_sheet_name(table.title or f"Table {idx + 1}")
            ws = wb.create_sheet(title=sheet_name)
            self._write_table(ws, table, result.validations)

        # Summary sheet
        if include_summary and result.tables:
            self._create_summary_sheet(wb, result)

        # Validation sheet
        if include_validation and result.validations:
            self._create_validation_sheet(wb, result.validations)

        # Metadata sheet
        self._create_metadata_sheet(wb, result)

        # Ensure at least one sheet
        if not wb.sheetnames:
            ws = wb.create_sheet("Empty")
            ws["A1"] = "No data extracted"

        # Save
        wb.save(str(output_path))
        wb.close()

        logger.info(f"Excel generated: {output_path}")
        return str(output_path)

    def _write_table(self, ws, table: Table, validations: list[ValidationResult]) -> None:
        """Write a table to a worksheet with full formatting."""
        if not table.cells:
            ws["A1"] = "No data"
            return

        # Build validation lookup
        val_lookup: dict[tuple[int, int], ValidationResult] = {}
        for v in validations:
            if v.row >= 0 and v.col >= 0:
                val_lookup[(v.row, v.col)] = v

        max_col_widths: dict[int, float] = {}
        confidence_threshold = Config.OCR_CONFIDENCE_THRESHOLD

        for row_idx, row_cells in enumerate(table.cells):
            for col_idx, cell in enumerate(row_cells):
                xl_row = row_idx + 1
                xl_col = col_idx + 1

                # Write value
                xl_cell = ws.cell(row=xl_row, column=xl_col)
                self._set_cell_value(xl_cell, cell)

                # Apply formatting
                if cell.is_header:
                    xl_cell.font = self.styles.header_font
                    xl_cell.fill = self.styles.header_fill
                    xl_cell.border = self.styles.header_border
                    xl_cell.alignment = self.styles.center_align
                else:
                    xl_cell.font = self.styles.bold_font if cell.font_bold else self.styles.data_font
                    xl_cell.border = self.styles.cell_border

                    # Alignment based on data type
                    if cell.data_type in ("number", "currency", "percentage"):
                        xl_cell.alignment = self.styles.right_align
                    elif cell.data_type == "date":
                        xl_cell.alignment = self.styles.center_align
                    else:
                        xl_cell.alignment = self.styles.left_align

                    # Number format
                    self._apply_number_format(xl_cell, cell)

                    # Low confidence highlighting
                    if cell.confidence < confidence_threshold:
                        xl_cell.fill = self.styles.low_confidence_fill
                        xl_cell.comment = self._make_comment(
                            f"Low confidence: {cell.confidence:.1%}\nOriginal: {cell.raw_value}"
                        )

                    # Validation error highlighting
                    val_key = (row_idx, col_idx)
                    if val_key in val_lookup:
                        v = val_lookup[val_key]
                        if v.severity == "error":
                            xl_cell.fill = self.styles.error_fill
                        elif v.severity == "warning":
                            xl_cell.fill = self.styles.warning_fill
                        xl_cell.comment = self._make_comment(v.message)

                    # Detect total rows
                    if col_idx == 0 and isinstance(cell.value, str):
                        if any(kw in cell.value.lower() for kw in ["total", "grand total", "sum", "net"]):
                            for c in range(len(row_cells)):
                                tc = ws.cell(row=xl_row, column=c + 1)
                                tc.fill = self.styles.total_fill
                                tc.font = self.styles.bold_font

                # Track column width
                value_str = str(cell.value) if cell.value else ""
                width = max(len(value_str) + 2, 8)
                max_col_widths[col_idx] = max(max_col_widths.get(col_idx, 8), min(width, 50))

        # Auto-width columns
        for col_idx, width in max_col_widths.items():
            col_letter = get_column_letter(col_idx + 1)
            ws.column_dimensions[col_letter].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        if table.cells:
            max_col = max(len(r) for r in table.cells)
            ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{len(table.cells)}"

    def _set_cell_value(self, xl_cell, cell: Cell) -> None:
        """Set cell value with proper type handling."""
        value = cell.value

        if value is None or value == "":
            xl_cell.value = None
        elif isinstance(value, datetime):
            xl_cell.value = value
        elif isinstance(value, (int, float)):
            xl_cell.value = value
        elif isinstance(value, str):
            # Prevent Excel from auto-converting
            if cell.data_type in ("gstin", "pan", "ifsc", "hsn", "sac"):
                xl_cell.value = value
                xl_cell.number_format = '@'  # Text format
            else:
                xl_cell.value = value
        else:
            xl_cell.value = str(value)

    def _apply_number_format(self, xl_cell, cell: Cell) -> None:
        """Apply appropriate number format to a cell."""
        if cell.format_string:
            xl_cell.number_format = cell.format_string
        elif cell.data_type == "currency":
            xl_cell.number_format = self.styles.currency_format_inr
        elif cell.data_type == "percentage":
            xl_cell.number_format = self.styles.percentage_format
        elif cell.data_type == "number":
            if isinstance(cell.value, float) and cell.value != int(cell.value):
                xl_cell.number_format = self.styles.number_format
            elif isinstance(cell.value, int):
                xl_cell.number_format = self.styles.integer_format
            else:
                xl_cell.number_format = self.styles.number_format
        elif cell.data_type == "date":
            xl_cell.number_format = self.styles.date_format
        elif cell.data_type in ("gstin", "pan", "ifsc", "hsn", "sac"):
            xl_cell.number_format = self.styles.text_format

    def _make_comment(self, text: str):
        """Create a cell comment."""
        from openpyxl.comments import Comment
        return Comment(text, "DocExcel Validator", width=250, height=100)

    def _create_summary_sheet(self, wb: Workbook, result: ProcessingResult) -> None:
        """Create a summary sheet with key statistics."""
        ws = wb.create_sheet(title="Summary", index=0)

        headers = ["Metric", "Value"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.styles.header_font
            cell.fill = self.styles.header_fill
            cell.border = self.styles.header_border
            cell.alignment = self.styles.center_align

        metrics = [
            ("File Name", result.filename),
            ("Document Type", result.document_type.value),
            ("Pages Processed", result.page_count),
            ("Tables Extracted", len(result.tables)),
            ("Total Rows", sum(t.row_count for t in result.tables)),
            ("OCR Engine", result.ocr_engine_used),
            ("Processing Time", f"{result.processing_time:.2f}s"),
            ("Validation Issues", len([v for v in result.validations if not v.is_valid])),
            ("Warnings", len(result.warnings)),
            ("Generated At", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ]

        for row_idx, (metric, value) in enumerate(metrics, 2):
            ws.cell(row=row_idx, column=1, value=metric).font = self.styles.bold_font
            ws.cell(row=row_idx, column=1).border = self.styles.cell_border
            ws.cell(row=row_idx, column=2, value=str(value)).border = self.styles.cell_border

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

        # Add numeric summaries if applicable
        row_offset = len(metrics) + 3

        for table in result.tables:
            if not table.cells or len(table.cells) < 2:
                continue

            ws.cell(row=row_offset, column=1, value=f"Table: {table.title or 'Data'}").font = self.styles.bold_font
            row_offset += 1

            for col_idx, header in enumerate(table.headers):
                # Calculate column sum for numeric columns
                total = 0.0
                count = 0
                for row in table.cells[1:]:
                    if col_idx < len(row):
                        cell = row[col_idx]
                        if isinstance(cell.value, (int, float)):
                            total += cell.value
                            count += 1

                if count > 0:
                    ws.cell(row=row_offset, column=1, value=f"  {header} (Sum)").border = self.styles.cell_border
                    ws.cell(row=row_offset, column=2, value=round(total, 2)).border = self.styles.cell_border
                    ws.cell(row=row_offset, column=2).number_format = self.styles.number_format
                    row_offset += 1

            row_offset += 1

    def _create_validation_sheet(self, wb: Workbook, validations: list[ValidationResult]) -> None:
        """Create a sheet listing all validation results."""
        ws = wb.create_sheet(title="Validation Report")

        headers = ["Row", "Column", "Field", "Check", "Status", "Expected", "Actual", "Message", "Severity"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.styles.header_font
            cell.fill = self.styles.header_fill
            cell.border = self.styles.header_border
            cell.alignment = self.styles.center_align

        for row_idx, v in enumerate(validations, 2):
            values = [
                v.row if v.row >= 0 else "",
                v.col if v.col >= 0 else "",
                v.field,
                v.check_type,
                "✓ Pass" if v.is_valid else "✗ Fail",
                str(v.expected_value) if v.expected_value else "",
                str(v.actual_value) if v.actual_value else "",
                v.message,
                v.severity.upper(),
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.styles.cell_border
                cell.font = self.styles.data_font

                if col_idx == 5:
                    if v.is_valid:
                        cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                    else:
                        cell.fill = self.styles.warning_fill

                if col_idx == 9:
                    if v.severity == "error":
                        cell.fill = self.styles.error_fill
                    elif v.severity == "warning":
                        cell.fill = self.styles.warning_fill

        # Auto-width
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(len(headers[col_idx - 1]) + 4, 12)

        ws.column_dimensions["H"].width = 60  # Message column
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(validations) + 1}"

    def _create_metadata_sheet(self, wb: Workbook, result: ProcessingResult) -> None:
        """Create metadata sheet with processing info."""
        ws = wb.create_sheet(title="Metadata")

        data = [
            ("Document ID", result.document_id),
            ("File Name", result.filename),
            ("File Hash", result.file_hash),
            ("File Category", result.file_category),
            ("Document Type", result.document_type.value),
            ("OCR Engine", result.ocr_engine_used),
            ("Pages", result.page_count),
            ("Tables", len(result.tables)),
            ("Processing Time", f"{result.processing_time:.2f}s"),
            ("Generated By", f"{Config.APP_NAME} v{Config.APP_VERSION}"),
            ("Generated At", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ]

        for row_idx, (key, value) in enumerate(data, 1):
            ws.cell(row=row_idx, column=1, value=key).font = self.styles.bold_font
            ws.cell(row=row_idx, column=2, value=str(value))

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 50

    def _safe_sheet_name(self, name: str) -> str:
        """Make a string safe for use as an Excel sheet name."""
        # Remove invalid characters
        invalid_chars = r'[]:*?/\\'
        safe = ''.join(c for c in name if c not in invalid_chars)
        # Limit to 31 characters
        return safe[:31] if safe else "Sheet"
