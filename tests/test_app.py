"""
Test suite for the Universal Document to Excel Converter.
Covers: utilities, models, validators, smart detection, excel generation.
"""
import os
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import pytest

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


# ═══════════════════════════════════════════════════════════════
# Utility Tests
# ═══════════════════════════════════════════════════════════════

class TestFormatDetection:
    """Test data type detection and number parsing."""

    def test_detect_plain_number(self):
        from backend.utils import detect_data_type
        dt, val, _ = detect_data_type("12345")
        assert dt == "number"
        assert val == 12345

    def test_detect_decimal(self):
        from backend.utils import detect_data_type
        dt, val, _ = detect_data_type("1234.56")
        assert dt == "number"
        assert val == 1234.56

    def test_detect_indian_number(self):
        from backend.utils import detect_data_type
        dt, val, _ = detect_data_type("1,23,456.78")
        assert dt == "number"
        assert val == 123456.78

    def test_detect_international_number(self):
        from backend.utils import detect_data_type
        dt, val, _ = detect_data_type("123,456.78")
        assert dt == "number"
        assert val == 123456.78

    def test_detect_currency_inr(self):
        from backend.utils import detect_data_type
        dt, val, fmt = detect_data_type("₹1,23,456.00")
        assert dt == "currency"
        assert val == 123456.0

    def test_detect_currency_dollar(self):
        from backend.utils import detect_data_type
        dt, val, _ = detect_data_type("$1,234.56")
        assert dt == "currency"
        assert val == 1234.56

    def test_detect_percentage(self):
        from backend.utils import detect_data_type
        dt, val, fmt = detect_data_type("18%")
        assert dt == "percentage"
        assert abs(val - 0.18) < 0.001
        assert fmt == "0.00%"

    def test_detect_negative_parenthetical(self):
        from backend.utils import detect_data_type
        dt, val, _ = detect_data_type("(1,234.56)")
        assert dt == "number"
        assert val == -1234.56

    def test_detect_gstin(self):
        from backend.utils import detect_data_type
        dt, val, _ = detect_data_type("27AABCU9603R1ZM")
        assert dt == "gstin"
        assert val == "27AABCU9603R1ZM"

    def test_detect_pan(self):
        from backend.utils import detect_data_type
        dt, val, _ = detect_data_type("AABCU9603R")
        assert dt == "pan"
        assert val == "AABCU9603R"

    def test_detect_ifsc(self):
        from backend.utils import detect_data_type
        dt, val, _ = detect_data_type("SBIN0001234")
        assert dt == "ifsc"

    def test_detect_date_ddmmyyyy(self):
        from backend.utils import detect_data_type
        dt, val, _ = detect_data_type("15/03/2024")
        assert dt == "date"

    def test_detect_text(self):
        from backend.utils import detect_data_type
        dt, val, _ = detect_data_type("Office Supplies")
        assert dt == "text"
        assert val == "Office Supplies"

    def test_detect_empty(self):
        from backend.utils import detect_data_type
        dt, val, _ = detect_data_type("")
        assert dt == "text"
        assert val == ""

    def test_detect_hsn_code(self):
        from backend.utils import detect_data_type
        dt, val, _ = detect_data_type("84714190")
        assert dt == "hsn"

    def test_detect_sac_code(self):
        from backend.utils import detect_data_type
        dt, val, _ = detect_data_type("998311")
        assert dt == "sac"


class TestNumberParsing:
    """Test number parsing with various formats."""

    def test_parse_plain(self):
        from backend.utils import parse_number
        assert parse_number("1234") == 1234.0

    def test_parse_commas(self):
        from backend.utils import parse_number
        assert parse_number("1,234,567.89") == 1234567.89

    def test_parse_indian(self):
        from backend.utils import parse_number
        assert parse_number("12,34,567.89") == 1234567.89

    def test_parse_currency_symbol(self):
        from backend.utils import parse_number
        assert parse_number("₹1,234.00") == 1234.0

    def test_parse_parenthetical_negative(self):
        from backend.utils import parse_number
        assert parse_number("(5,678.90)") == -5678.90

    def test_parse_debit_notation(self):
        from backend.utils import parse_number
        result = parse_number("1,234.00 Dr")
        assert result == -1234.0

    def test_parse_empty(self):
        from backend.utils import parse_number
        assert parse_number("") is None

    def test_parse_non_numeric(self):
        from backend.utils import parse_number
        assert parse_number("hello") is None


class TestDateParsing:
    """Test date parsing with various formats."""

    def test_parse_dd_mm_yyyy(self):
        from backend.utils import parse_date
        result = parse_date("15/03/2024")
        assert result == datetime(2024, 3, 15)

    def test_parse_dd_mm_yyyy_dash(self):
        from backend.utils import parse_date
        result = parse_date("15-03-2024")
        assert result == datetime(2024, 3, 15)

    def test_parse_yyyy_mm_dd(self):
        from backend.utils import parse_date
        result = parse_date("2024-03-15")
        assert result == datetime(2024, 3, 15)

    def test_parse_invalid(self):
        from backend.utils import parse_date
        assert parse_date("not a date") is None


class TestIndianNumber:
    """Test Indian number formatting."""

    def test_format_lakhs(self):
        from backend.utils import format_indian_number
        result = format_indian_number(1234567.89)
        assert result == "12,34,567.89"

    def test_format_crores(self):
        from backend.utils import format_indian_number
        result = format_indian_number(12345678.00)
        assert result == "1,23,45,678.00"

    def test_format_small(self):
        from backend.utils import format_indian_number
        result = format_indian_number(999.99)
        assert result == "999.99"


class TestFileUtilities:
    """Test file utility functions."""

    def test_safe_filename(self):
        from backend.utils import safe_filename
        assert safe_filename("my/file:name.pdf") == "my_file_name.pdf"

    def test_safe_filename_empty(self):
        from backend.utils import safe_filename
        assert safe_filename("...") == "unnamed_file"

    def test_get_extension(self):
        from backend.utils import get_file_extension
        assert get_file_extension("test.PDF") == "pdf"
        assert get_file_extension("test.xlsx") == "xlsx"

    def test_human_readable_size(self):
        from backend.utils import get_human_readable_size
        assert "KB" in get_human_readable_size(1500)
        assert "MB" in get_human_readable_size(1500000)


# ═══════════════════════════════════════════════════════════════
# Model Tests
# ═══════════════════════════════════════════════════════════════

class TestModels:
    """Test data model serialization and behavior."""

    def test_cell_to_dict(self):
        from backend.models import Cell
        cell = Cell(value="Test", confidence=0.95, data_type="text", row=0, col=0)
        d = cell.to_dict()
        assert d["value"] == "Test"
        assert d["confidence"] == 0.95

    def test_table_to_dict(self):
        from backend.models import Cell, Table
        table = Table(title="Test Table")
        table.cells.append([Cell(value="Header", is_header=True)])
        table.cells.append([Cell(value="Data")])
        d = table.to_dict()
        assert d["title"] == "Test Table"
        assert d["num_rows"] == 2
        assert len(d["cells"]) == 2

    def test_result_lifecycle(self):
        from backend.models import ProcessingResult, ProcessingStatus
        result = ProcessingResult()
        assert result.status == ProcessingStatus.QUEUED

        result.mark_completed()
        assert result.status == ProcessingStatus.COMPLETED
        assert result.processing_time > 0

    def test_result_failure(self):
        from backend.models import ProcessingResult, ProcessingStatus
        result = ProcessingResult()
        result.mark_failed("Test error")
        assert result.status == ProcessingStatus.FAILED
        assert result.error_message == "Test error"


# ═══════════════════════════════════════════════════════════════
# Validator Tests
# ═══════════════════════════════════════════════════════════════

class TestValidators:
    """Test financial validation rules."""

    def test_valid_gstin(self):
        from backend.validators import GSTIN_PATTERN
        assert GSTIN_PATTERN.match("27AABCU9603R1ZM")
        assert not GSTIN_PATTERN.match("INVALID")

    def test_valid_pan(self):
        from backend.validators import PAN_PATTERN
        assert PAN_PATTERN.match("AABCU9603R")
        assert not PAN_PATTERN.match("12345")

    def test_valid_ifsc(self):
        from backend.validators import IFSC_PATTERN
        assert IFSC_PATTERN.match("SBIN0001234")
        assert not IFSC_PATTERN.match("INVALID")

    def test_invalid_state_code(self):
        from backend.validators import VALID_STATE_CODES
        assert "27" in VALID_STATE_CODES  # Maharashtra
        assert "99" not in VALID_STATE_CODES

    def test_validate_table_arithmetic(self):
        from backend.models import Cell, Table
        from backend.validators import FinancialValidator

        table = Table()
        # Header
        table.cells.append([
            Cell(value="Qty", is_header=True),
            Cell(value="Rate", is_header=True),
            Cell(value="Amount", is_header=True),
        ])
        table.headers = ["Qty", "Rate", "Amount"]
        # Data with correct arithmetic
        table.cells.append([
            Cell(value=10, data_type="number"),
            Cell(value=100, data_type="number"),
            Cell(value=1000, data_type="number"),
        ])

        validator = FinancialValidator()
        results = validator.validate_table(table)
        # Should pass arithmetic check
        arithmetic_errors = [r for r in results if r.check_type == "arithmetic" and not r.is_valid]
        assert len(arithmetic_errors) == 0

    def test_validate_gst_cgst_sgst(self):
        from backend.models import Cell, Table
        from backend.validators import FinancialValidator

        table = Table()
        table.cells.append([
            Cell(value="CGST", is_header=True),
            Cell(value="SGST", is_header=True),
        ])
        table.headers = ["CGST", "SGST"]
        # Mismatched CGST and SGST
        table.cells.append([
            Cell(value=900, data_type="number"),
            Cell(value=450, data_type="number"),
        ])

        validator = FinancialValidator()
        results = validator.validate_table(table)
        gst_errors = [r for r in results if r.check_type == "gst" and not r.is_valid]
        assert len(gst_errors) > 0


# ═══════════════════════════════════════════════════════════════
# Smart Detection Tests
# ═══════════════════════════════════════════════════════════════

class TestSmartDetection:
    """Test column detection and document classification."""

    def test_column_detection(self):
        from backend.models import Cell, Table
        from backend.smart_detection import SmartColumnDetector

        table = Table()
        table.cells.append([
            Cell(value="Inv No", is_header=True),
            Cell(value="Dt", is_header=True),
            Cell(value="Amt", is_header=True),
        ])
        table.headers = ["Inv No", "Dt", "Amt"]

        detector = SmartColumnDetector()
        result = detector.detect_and_rename(table)
        # Should rename headers
        assert result.headers[0] == "Invoice No"
        assert result.headers[1] == "Date"
        assert result.headers[2] == "Amount"

    def test_document_classification(self):
        from backend.smart_detection import DocumentClassifier
        from backend.models import DocumentType

        classifier = DocumentClassifier()

        # Bank statement
        result = classifier.classify(text="Bank Statement Account Summary")
        assert result == DocumentType.BANK_STATEMENT

        # Invoice
        result = classifier.classify(text="Tax Invoice GST")
        assert result == DocumentType.INVOICE


# ═══════════════════════════════════════════════════════════════
# Excel Generation Tests
# ═══════════════════════════════════════════════════════════════

class TestExcelGeneration:
    """Test Excel workbook generation."""

    def test_generate_basic_excel(self):
        from backend.models import Cell, Table, ProcessingResult
        from backend.excel import ExcelGenerator

        result = ProcessingResult()
        result.filename = "test.pdf"

        table = Table(title="Test Data")
        table.cells.append([
            Cell(value="Name", is_header=True, font_bold=True),
            Cell(value="Amount", is_header=True, font_bold=True),
        ])
        table.headers = ["Name", "Amount"]
        table.cells.append([
            Cell(value="Item A", data_type="text"),
            Cell(value=1500.50, data_type="number", format_string="#,##0.00"),
        ])
        table.cells.append([
            Cell(value="Item B", data_type="text"),
            Cell(value=2300.00, data_type="number", format_string="#,##0.00"),
        ])

        result.tables.append(table)

        generator = ExcelGenerator()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            path = generator.generate(result, output_path=output_path)
            assert os.path.exists(path)
            assert os.path.getsize(path) > 0

            # Verify content
            from openpyxl import load_workbook
            wb = load_workbook(path)
            assert "Test Data" in wb.sheetnames
            assert "Summary" in wb.sheetnames
            assert "Metadata" in wb.sheetnames

            ws = wb["Test Data"]
            assert ws.cell(1, 1).value == "Name"
            assert ws.cell(2, 1).value == "Item A"
            assert ws.cell(2, 2).value == 1500.50
            wb.close()
        finally:
            os.unlink(output_path)

    def test_generate_with_validations(self):
        from backend.models import Cell, Table, ProcessingResult, ValidationResult
        from backend.excel import ExcelGenerator

        result = ProcessingResult()
        result.filename = "test_validation.pdf"

        table = Table(title="Invoice Data")
        table.cells.append([
            Cell(value="GSTIN", is_header=True),
            Cell(value="Amount", is_header=True),
        ])
        table.headers = ["GSTIN", "Amount"]
        table.cells.append([
            Cell(value="INVALID_GSTIN", data_type="text"),
            Cell(value=1000, data_type="number"),
        ])
        result.tables.append(table)

        result.validations.append(ValidationResult(
            field="GSTIN", check_type="format", is_valid=False,
            actual_value="INVALID_GSTIN", message="Invalid GSTIN format",
            severity="error", row=1, col=0,
        ))

        generator = ExcelGenerator()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            path = generator.generate(result, output_path=output_path)
            assert os.path.exists(path)

            from openpyxl import load_workbook
            wb = load_workbook(path)
            assert "Validation Report" in wb.sheetnames
            wb.close()
        finally:
            os.unlink(output_path)


# ═══════════════════════════════════════════════════════════════
# Config Tests
# ═══════════════════════════════════════════════════════════════

class TestConfig:
    """Test configuration module."""

    def test_file_categories(self):
        from config import Config
        assert Config.get_file_category("pdf") == "pdf"
        assert Config.get_file_category("jpg") == "image"
        assert Config.get_file_category("xlsx") == "excel"
        assert Config.get_file_category("docx") == "word"
        assert Config.get_file_category("zip") == "archive"
        assert Config.get_file_category("xyz") == "unknown"

    def test_ensure_directories(self):
        from config import Config
        Config.ensure_directories()
        assert Config.TEMP_DIR.exists()
        assert Config.LOG_DIR.exists()


# ═══════════════════════════════════════════════════════════════
# Flask App Tests
# ═══════════════════════════════════════════════════════════════

class TestFlaskApp:
    """Test Flask application endpoints."""

    @pytest.fixture
    def client(self):
        from app import create_app
        app = create_app()
        app.config["TESTING"] = True
        with app.test_client() as client:
            yield client

    def test_health_check(self, client):
        response = client.get("/api/health")
        assert response.status_code == 200
        data = response.get_json()
        assert data["status"] == "healthy"

    def test_index_page(self, client):
        response = client.get("/")
        assert response.status_code == 200
        assert b"DocExcel" in response.data

    def test_get_settings(self, client):
        response = client.get("/api/settings")
        assert response.status_code == 200
        data = response.get_json()
        assert "confidence_threshold" in data
        assert "supported_formats" in data

    def test_upload_no_files(self, client):
        response = client.post("/api/upload")
        assert response.status_code == 400

    def test_job_not_found(self, client):
        response = client.get("/api/job/nonexistent")
        assert response.status_code == 404


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
